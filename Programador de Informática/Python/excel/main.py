import openpyxl
import string
from pathlib import Path
xlsx_file = Path('C:/repositorio', 'python_auto_mail.xlsx')
print(xlsx_file)
wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj['Folha1']
print(sheet['a1'].value, sheet['b1'].value)
sheet['a3'].value = 'Daniel'
for n in range (1, 2):
    print(sheet.cell(row=n, column=1).value)
for n in range(1, 11):
    sheet[f'a{n}'].value = n
for n in range(1,11):
    print(sheet.cell(row=n, column=1).value)
    print(sheet[f'a{n}'].value)
